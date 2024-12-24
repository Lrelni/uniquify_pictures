import os
import sys
import time
import shutil
import hashlib
import operator
from multiprocessing import Pool

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font # type: ignore
from tqdm import tqdm # type: ignore

DEBUG = False
REPORT = False
DUREPORT = False
PLACE = False
INPUTPATH = "./uqinput"
OUTPUTPATH = "./uqoutput"

# taken directly from PIL.Image.registered_extensions()
EXTENSIONS = ['.blp', '.bmp', '.dib', '.bufr', '.cur', '.pcx', '.dcx', '.dds', 
'.ps', '.eps', '.fit', '.fits', '.fli', '.flc', '.ftc', '.ftu', '.gbr', '.gif', 
'.grib', '.h5', '.hdf', '.png', '.apng', '.jp2', '.j2k', '.jpc', '.jpf',
'.jpx', '.j2c', '.icns', '.ico', '.im', '.iim', '.jfif', '.jpe', '.jpg',
'.jpeg', '.mpg', '.mpeg', '.tif', '.tiff', '.mpo', '.msp', '.palm', '.pcd',
'.pdf', '.pxr', '.pbm', '.pgm', '.ppm', '.pnm', '.pfm', '.psd', '.qoi',
'.bw', '.rgb', '.rgba', '.sgi', '.ras', '.tga', '.icb', '.vda', '.vst', 
'.webp', '.wmf', '.emf', '.xbm', '.xpm']
MPLIMIT = 300

def scan_pictures(path):
    # scan path for all the pictures
    # /uqinput is used for this program

    raw_locations = []
    locations = []
    other = []

    # traverse folder and get all the files
    print("Scanning "+path+"...")
    for (root, dirs, file) in os.walk(path):
        for f in file:
            raw_locations.append(os.path.join(root, f))
    
    # check for non-pictures
    raw_length = len(raw_locations)
    print(str(raw_length) + " files found.")
    print("Keeping image files and ignoring other files...")

    for loc in tqdm(raw_locations):
        if check_image(loc):
            # if it is a picture
            locations.append(loc)
        else:
            other.append(loc)
    print(str(len(locations))+" pictures in "+path)

    return (locations, other)

def check_image(f):
    if os.path.splitext(f)[1] in EXTENSIONS:
        try:
            open(f, "rb")
            return True
        except:
            return False
    else:
        return False

def ikey(f):
    # hash images to make image keys for sorting
    return hashlib.file_digest(open(f, "rb"), "sha256").hexdigest()

def pkg_ikey(x):
    # simply return a tuple of the path and its ikey
    return (x, ikey(x))

def cache_ikey(data):
    # takes in a list of paths (should be check_imaged) 
    # and outputs them as [ (location, ikey), etc. ]
    print("Computing sorting hashes...")

    result = None;

    if len(data) > MPLIMIT:
        with Pool() as p:
            result = list(p.map(pkg_ikey, tqdm(data)))
    else:
        result = map(pkg_ikey, tqdm(data))

    return result

def report(sorted_data, tstamp, du=False):
    # takes in a list of paths sorted by image hashes and writes them to xlsx
    print("Generating xlsx " + ("duplicates " if du else "") + "report...")

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Filename")
    ws.cell(row=1, column=2, value="Path")
    ws.cell(row=1, column=3, value="Link")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 60

    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)

    if du:
        ws.cell(row=1, column=4, value="Note: Only duplicates in this file.")
        ws.column_dimensions['D'].width = 80
        ws['D1'].font = Font(bold=True)

    cur_row = 2
    len_data = len(sorted_data)

    def insert(path, nl=False):
        # insert an entry into the spreadsheet
        nonlocal cur_row, ws

        cur_row += 1 if nl else 0 #newline before entry if nl is true
        real = os.path.realpath(path)
        fp = os.path.split(real)
        ws.cell(row=cur_row, column=1, value=fp[1])
        ws.cell(row=cur_row, column=2, value=fp[0])
        ws.cell(row=cur_row, column=3, value='=HYPERLINK("'+real+'")')
        cur_row += 1

    if du:
        for index, entry in enumerate(tqdm(sorted_data)):
            if index == 0: # first element
                if entry[1] == sorted_data[index + 1][1]:
                    insert(entry[0])
            elif index == len_data - 1: # last element
                if entry[1] == sorted_data[index-1]:
                    insert(entry[0])
            else: # handle all in between
                if sorted_data[index-1][1] == entry[1]:
                    # confirmed to be a duplicate
                    insert(entry[0])
                else:
                    # either start of new chain or unique
                    if sorted_data[index+1][1] == entry[1]:
                        insert(entry[0], nl=True) # new chain
                    # if not equal to next element either,
                    # then it is unique and should not be inserted
    else:
        for index, entry in enumerate(tqdm(sorted_data)):
            if index == 0: # first element 
                insert(entry[0])
            else:
                if entry[1] == sorted_data[index-1][1]:
                    # same as last
                    insert(entry[0])
                else:
                    insert(entry[0], nl=True)

    report_name = os.path.join(OUTPUTPATH, "uqreport_"+("duplicates_" if du else "")+str(tstamp)+".xlsx")
    wb.save(report_name)
    print("Report created at "+report_name)

def sorted_cleaner(sorted_data):
    # takes in a list of paths sorted by image hashes and then uniquifies them
    print("Uniquifying pictures...")
    unique = []
    for index, entry in enumerate(tqdm(sorted_data)):
        if index == 0:
            unique.append(entry)
        elif sorted_data[index-1][1] != entry[1]:
            unique.append(entry)
    return unique
 
def clean(folder):
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                # if file, remove
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                # if dir, also remove
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))
        
def main():
    print("\n\u2588 uq.py: Remove duplicate pictures. \u2588\n")

    if PLACE or REPORT or DUREPORT:
        os.makedirs(OUTPUTPATH, exist_ok=True)
        if len(os.listdir(OUTPUTPATH)) > 0: 
            # only clear if there are files
            istr = "Files already exist in the output path.\
 Should the files be cleared? y/n: "
            should_clear = input(istr)
            should_clear = True if should_clear == "Y" or should_clear == "y" else False
            if should_clear:
                print("Cleaning "+OUTPUTPATH+"...")
                clean(OUTPUTPATH)
            else:
                print("Output directory is not cleared. Exiting.")
                return

    scanned = scan_pictures(INPUTPATH) # tuple (pictures, non-pictures)
    print("")
    sloc = sorted(cache_ikey(scanned[0]),\
        key=operator.itemgetter(1))

    timestamp = int(1000 * time.time())
    print("")
    if REPORT:
        report(sloc, timestamp)
    print("")
    if DUREPORT:
        report(sloc, timestamp, du=True)
    print("")
    if PLACE:
        unique = sorted_cleaner(sloc)
        print("Placing unique pictures into "+OUTPUTPATH)

        for entry in tqdm(unique):
            rel_path = os.path.relpath(entry[0], INPUTPATH)
            dest = os.path.join(OUTPUTPATH, rel_path)
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            shutil.copy2(entry[0], dest)

        print("Placing other files (non-pictures) into "+OUTPUTPATH)

        for file in tqdm(scanned[1]): # keep all non-pictures
            rel_path = os.path.relpath(file, INPUTPATH)
            dest = os.path.join(OUTPUTPATH, rel_path)
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            shutil.copy2(file, dest)


def debug():
    pass

if __name__ == "__main__":
    if "-r" in sys.argv:
        REPORT = True

    if "-p" in sys.argv:
        PLACE = True
    
    if "-d" in sys.argv:
        DUREPORT = True
    
    if "-i" in sys.argv:
        INPUTPATH = sys.argv[sys.argv.index("-i")+1]
    else:
        print("No input directory specified with -i; defaulting to ./uqinput")
    
    if "-o" in sys.argv:
        OUTPUTPATH = sys.argv[sys.argv.index("-o")+1]
    else:
        print("No output directory specified with -o; defaulting to ./uqoutput")
    
    if (not REPORT) and (not PLACE):
        print("Warning: no recognized arguments specified.")
        print("Use -r to generate a report of all files.")
        print("Use -d to generate a report of duplicates.")
        print("Use -p to copy unique pictures into output.")
        print("Use -i [input path] to specify an input path.")
        print("Use -o [output path] to specify an output path.")

    if DEBUG:
        print("Running in debug mode.")
        debug()
    else:
        main()
